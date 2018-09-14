import os
from abc import ABC, abstractmethod, ABCMeta

import pygsheets
from openpyxl import Workbook, load_workbook

from cellbase.celltable import LocalCelltable, GoogleCelltable


class Cellbase(ABC):
    """
    Cellbase is equivalent to Workbook that consist of Worksheets
    """
    DEFAULT_FILENAME = 'cellbase.xlsx'

    def __init__(self):
        self._path = ''
        self._filename = Cellbase.DEFAULT_FILENAME
        self._workbook = None
        self.schemas = {}
        self.celltables = {}

    @property
    def path(self):
        """ Path to the workbook """
        return self._path

    @property
    def filename(self):
        """ Name of workbook """
        return self._filename

    @property
    def dir(self):
        """ Full directory of workbook(path/filename) """
        return os.path.join(self.path, self.filename)

    def load(self, path, filename, raise_err=False):
        """
        Load workbook from given dir(path/filename) to memory.

        By default, it automatically create workbook if not found & no error will be raised.
        Set raise_err to True to change this behavior.
        """
        self._path = path
        self._filename = filename
        self._on_load(raise_err)
        return self

    def create_if_none(self, name):
        """
        Create worksheet and add to celltables if there's no such worksheet.
        """
        if name not in self.celltables:
            if name not in self.schemas:
                raise ValueError(
                    "Trying to create Celltable '%s' without schema " % name)
            self.celltables[name] = self._on_create(name)

    def drop(self, name):
        """ Delete specified worksheet. """
        self._on_drop(name)
        self.celltables.pop(name)

    def save_as(self, path, filename, overwrite=False):
        """
        Save workbook to dir(path/filename).

        FileExistsError will be raised if file exists. Set overwrite to True to allow overwriting.
        """
        if os.path.exists(filename) and not overwrite:
            raise FileExistsError("%s already exists, set overwrite=True if this is expected.")
        self._on_save(path, filename)

    def save(self):
        """ Save workbook to the filename specified in load, overwrite if file exist. """
        self.save_as(self.path, self.filename, overwrite=True)

    @abstractmethod
    def _on_load(self, raise_err):
        """ When actual loading perform """
        pass

    @abstractmethod
    def _on_create(self, name):
        """ When creating new workbook, expected to return new Celltable instance """
        pass

    @abstractmethod
    def _on_drop(self, worksheet):
        """ When deleting celltable/worksheet """
        pass

    @abstractmethod
    def _on_save(self, path, filename):
        """
        When actual saving performed.

        Should be saved to given path & filename instead of self.path & self.filename, as save_as is the base method of
        save, so there's no guarantee that self.path & self.filename are the desire directory when saving.
        """
        pass

    def register(self, schema):
        """
        Register schema of worksheet to deal with, only required for newly created worksheet

        Example::

            {'TABLE_NAME': ['COL_NAME_1', 'COL_NAME_2']}
        """
        self.schemas.update(schema)

    def query(self, name, where=None):
        """
        Return data from Celltable with specified worksheet, that match the conditions.

        Return all data if where omitted.

        Example::

            where = {'id': 1, 'name': 'jp'}

        :return:
            List of dict that store value corresponding to the row_idx & column id.
            For example, [{"row_idx": 2, "id": 1, "name": "jp1"}, {"row_idx": 3, "id": 2, "name": "jp2"}]
        """
        self.create_if_none(name)
        return self.celltables[name].query(where=where)

    def insert(self, name, value_in_dict):
        """
        Insert new row to the worksheet

        Example::

            value_in_dict = {"id": 1, "name": "jp1"}
        :return: row_idx of new row
        """
        self.create_if_none(name)
        return self.celltables[name].insert(value_in_dict)

    def update(self, name, value_in_dict, where=None):
        """
        Update row(s) that match the condition.

        If row_idx is given in value_in_dict, where will be ignored and only the exact row will be updated.

        Example::

            where = {'id': 1, 'name': 'jp'}.
        :return: Number of updated rows
        """
        self.create_if_none(name)
        return self.celltables[name].update(value_in_dict, where=where)

    def delete(self, name, where=None):
        """
        Delete row(s) that match conditions.

        Delete all if where omitted.

        Example::

            where = {'id': 1, 'name': 'jp'}.
        :return: Number of deleted rows
        """
        self.create_if_none(name)
        return self.celltables[name].delete(where=where)

    def traverse(self, name, fn, where=None, select=None):
        """
        Access cells directly from rows where conditions matched.

        Select all column if select omitted

        Example::

            fn = lambda cell: cell.fill = PatternFill(fill_type="solid", fgColor="00FFFF00").

            select = ['id']  # only column under "id" will be accessed
        :return: Number of traversed rows
        """
        self.create_if_none(name)
        return self.celltables[name].traverse(fn, where=where, select=select)

    def format(self, name, formatter, where=None, select=None):
        """
        Convenience method that built on top of traverse to format cell(s).

        formatter can be :class:`cellbase.formatter.CellFormatter` or dict

        :return: Number of formatted rows
        """
        self.create_if_none(name)
        return self.celltables[name].format(formatter, where=where, select=select)

    def __len__(self):
        """ Return numbers of worksheet """
        return len(self.celltables)

    def __getitem__(self, name):
        """ Get Celltable with worksheet name """
        self.create_if_none(name)
        return self.celltables[name]

    def __setitem__(self, name, celltable):
        """ Celltable must be created by Cellbase, AssertionError raised when attempt to assign """
        raise AssertionError("Celltable must be created by Cellbase")

    def __delitem__(self, name):
        """ Drop worksheet  """
        self.drop(name)

    def __contains__(self, name):
        """ Check if worksheet exists """
        return name in self.celltables


class LocalCellbase(Cellbase):
    """ Cellbase that handle local workbook with openpyxl """

    def __init__(self):
        super().__init__()
        self._workbook = Workbook()

    def _on_load(self, raise_err):
        if os.path.exists(self.dir):
            self._workbook = load_workbook(self.dir)
        elif raise_err:
            raise FileNotFoundError("No workbook found at %s" % self.dir)
        else:
            self._workbook = Workbook()
        for worksheet in self._workbook.worksheets:
            self.celltables[worksheet.title] = LocalCelltable(worksheet)

    def _on_create(self, name):
        worksheet = self._workbook.create_sheet(title=name, index=0)
        worksheet.append(self.schemas[name])
        return LocalCelltable(worksheet)

    def _on_drop(self, name):
        worksheet_to_drop = self.celltables[name].worksheet
        # Workbook must contain at least 1 visible sheet
        visible_sheets = [worksheet for worksheet in self._workbook.worksheets
                          if worksheet.sheet_state == 'visible']
        if len(visible_sheets) == 1 and visible_sheets[0] is worksheet_to_drop:
            self._workbook.create_sheet()
        self._workbook.remove(worksheet_to_drop)

    def _on_save(self, path, filename):
        self._workbook.save(os.path.join(path, filename))


class RemoteCellbase(Cellbase, metaclass=ABCMeta):
    """ Cellbase that handle remote workbook """

    def __init__(self, **kwargs):
        super().__init__()
        unexpected_attrs = [attr for attr in kwargs if attr not in self.attrs()]
        if unexpected_attrs:
            raise AttributeError("Unexpected attribute%s, expecting%s only" %
                                 (unexpected_attrs, GoogleCellbase.ATTRIBUTES))
        else:
            self.__dict__.update(kwargs)

    @abstractmethod
    def attrs(self):
        """ Possible attributes required """
        pass

    def __getattr__(self, attr):
        try:
            return self.__dict__[attr]
        except KeyError:
            if attr not in self.attrs():
                raise AttributeError("Unexpected attribute %s, expecting%s only" % (attr, GoogleCellbase.ATTRIBUTES))
            else:
                self.__dict__[attr] = None
                return None


class GoogleCellbase(RemoteCellbase):
    """ Cellbase that handle workbook from Google Drive through Google Sheet API """
    ATTRIBUTES = ('client_secret', 'service_account_file', 'credentials_directory')

    def __init__(self, export_path='', export_format=pygsheets.ExportType.CSV, **kwargs):
        """
        Both export_path & export_format are used when saving, as GoogleCellbase's save method behave differently which
        export google spreadsheet to local.

        :param kwargs: See GoogleCellbase.ATTRIBUTES for possible arguments
        """
        super().__init__(**kwargs)
        self.export_path = export_path
        self.export_format = export_format

    def attrs(self):
        return GoogleCellbase.ATTRIBUTES

    def load(self, path, filename, raise_err=False):
        """
        Path can be `folder id <https://developers.google.com/drive/api/v3/folder/>`_ in Google Drive or empty string
        where workbook will be created in root folder of Google Drive by default.
        """
        super().load(path, filename, raise_err)

    def save(self):
        """
        As Google Spreadsheet does not required saving, save method is used to export Google Spreadsheet to export_path
        and same filename as loaded spreadsheet.
        """
        self.save_as(self.export_path, self.filename, True)

    def save_as(self, path, filename, overwrite=False):
        """
        As Google Spreadsheet does not required saving, save_as method is used to export Google Spreadsheet to given
        path and filename.

        FileExistsError will be raised if file exists. Set overwrite to True to allow overwriting.
        """
        super().save_as(path, filename, overwrite)

    def _on_load(self, raise_err):
        client = pygsheets.authorize(self.client_secret or 'client_secret.json', self.service_account_file,
                                     self.credentials_directory)
        try:
            self._workbook = client.open(self._filename)
        except pygsheets.SpreadsheetNotFound:
            if raise_err:
                raise FileNotFoundError("No workbook found at %s" % self.filename)
            else:
                self._workbook = client.create(self._filename, folder=self.path or 'root')
        for worksheet in self._workbook.worksheets():
            self.celltables[worksheet.title] = GoogleCelltable(worksheet, worksheet.title in self.schemas)

    def _on_create(self, name):
        worksheet = self._workbook.add_worksheet(title=name, index=0)
        worksheet.update_row(1, self.schemas[name])
        return GoogleCelltable(worksheet)

    def _on_drop(self, name):
        worksheet_to_drop = self._workbook.worksheet_by_title(name)
        # Workbook must contain at least 1 visible sheet
        visible_sheets = [worksheet for worksheet in self._workbook.worksheets()
                          if not worksheet.jsonSheet['properties'].get('hidden', False)]
        if len(visible_sheets) == 1 and visible_sheets[0] is worksheet_to_drop:
            titles = [worksheet.title for worksheet in self._workbook.worksheets()]
            self._workbook.add_worksheet(new_worksheet_title(titles))
        self._workbook.del_worksheet(worksheet_to_drop)

    def _on_save(self, path, filename):
        self._workbook.export(self.export_format, path, filename)


def new_worksheet_title(titles, counter=0, name='Sheet'):
    """ Simple helper to recursively generate new title """
    new_name = '%s%s' % (name, counter or '')
    if new_name in titles:
        return new_worksheet_title(titles, counter + 1, name)
    else:
        return new_name
