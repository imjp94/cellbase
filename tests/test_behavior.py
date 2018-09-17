import pytest
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.styles.numbers import FORMAT_TEXT

from cellbase import LocalCellbase, LocalCelltable
from cellbase.formatter import LocalCellFormatter
from cellbase.helper import DAO, Entity


class TestBehavior:
    @pytest.fixture
    def cellbase(self):
        cellbase = LocalCellbase().load("res", "not_exist.xlsx")
        cellbase.register(SimpleDAO.on_create())
        return cellbase

    @pytest.fixture
    def dao(self, cellbase):
        return SimpleDAO(cellbase)


class TestCelltableBehavior(TestBehavior):
    def test_insert(self, dao):
        assert dao.celltable.size == 0
        simple = populate(dao)[0]
        assert dao.celltable.size == 1
        assert simple == dao.query()[0]

    def test_query_all(self, dao):
        num_row = 5
        populate(dao, num_row)
        assert num_row == len(dao.query())

    def test_query_with_callable(self, dao):
        num_row = 5
        populate(dao, num_row)  # 5 rows inserted to row index 2, 3, 4, 5, 6
        simples_4_to_6 = dao[lambda row_idx: 4 <= row_idx <= 6]
        assert len(simples_4_to_6) == 3
        for i in range(4, 7):
            assert simples_4_to_6[i - 4].row_idx == i

    def test_update(self, dao):
        simple = populate(dao)[0]
        simple_to_update = Simple(id=3, name="updated_simple")
        simple_to_update.row_idx = simple.row_idx
        assert simple_to_update != dao.query({DAO.COL_ROW_IDX: simple.row_idx})[0]
        dao.update(simple_to_update, {DAO.COL_ROW_IDX: 2})
        assert simple_to_update == dao.query({DAO.COL_ROW_IDX: simple.row_idx})[0]
        assert simple != simple_to_update

    def test_delete(self, dao):
        simple = populate(dao)[0]
        deleted_count = dao.delete({DAO.COL_ROW_IDX: simple.row_idx})
        assert deleted_count == 1
        assert dao.query({DAO.COL_ROW_IDX: simple.row_idx}) == []

    def test_format(self, dao):
        simple_formatted = Simple(id=5, name="simple_formatted")
        simple_not_formatted = Simple(id=6, name="simple_not_formatted")
        dao.insert(simple_formatted)
        dao.insert(simple_not_formatted)
        font = Font(name='Arial')
        fill = PatternFill(fill_type="solid", fgColor="00FFFF00")
        border = Border(top=Side(style="thin"))
        alignment = Alignment(horizontal="left")
        number_format = FORMAT_TEXT
        protection = Protection(hidden=True)
        formatter = LocalCellFormatter(
            font=font, fill=fill, border=border,
            number_format=number_format, protection=protection,
            alignment=alignment)
        dao.format({'font': font, 'fill': fill, 'border': border, 'number_format': number_format,
                    'protection': protection, 'alignment': alignment}, {DAO.COL_ROW_IDX: simple_formatted.row_idx})

        def format_match(cell, fmtter):
            assert all_format(cell, fmtter)

        def format_not_match(cell, fmtter):
            assert not all_format(cell, fmtter)

        dao.traverse(lambda cell: format_match(cell, formatter), {DAO.COL_ROW_IDX: simple_formatted.row_idx})
        dao.traverse(lambda cell: format_not_match(cell, formatter), {DAO.COL_ROW_IDX: simple_not_formatted.row_idx})


class TestCellbaseBehavior(TestBehavior):
    @pytest.fixture
    def cellbase(self):
        cellbase = super().cellbase()
        cellbase.create_if_none(SimpleDAO.TABLE_NAME)
        return cellbase

    def test_contains(self, cellbase):
        assert SimpleDAO.TABLE_NAME in cellbase

    def test_getitem(self, cellbase):
        assert cellbase[SimpleDAO.TABLE_NAME] == cellbase.celltables[SimpleDAO.TABLE_NAME]

    def test_setitem(self, cellbase):
        with pytest.raises(AssertionError):
            cellbase[SimpleDAO.TABLE_NAME] = LocalCelltable(cellbase._workbook.create_sheet(title=SimpleDAO.TABLE_NAME))

    def test_delitem(self, cellbase):
        del cellbase.celltables[SimpleDAO.TABLE_NAME]
        assert SimpleDAO.TABLE_NAME not in cellbase

    def test_len(self, cellbase):
        assert len(cellbase) == 2  # Default worksheet + Simple worksheet


class TestHelperBehavior(TestBehavior):
    def test_contains(self, dao):
        simple = populate(dao)[0]
        assert simple.row_idx in dao

    def test_getitem(self, dao):
        simple = populate(dao)[0]
        assert simple == dao[simple.row_idx]

    def test_getitem_with_callable(self, dao):
        simple = populate(dao)[0]
        assert simple == dao[lambda row_idx: row_idx == simple.row_idx][0]

    def test_setitem(self, dao):
        simple = populate(dao)[0]
        assert dao[simple.row_idx].name == simple.name
        new_name = "updated_simple"
        simple.name = new_name
        dao[simple.row_idx] = simple
        assert dao[simple.row_idx].name == new_name

    def test_setitem_with_callable(self, dao):
        simple_new = Simple(id=0, name='new_simple')
        with pytest.warns(UserWarning):
            dao[lambda row_idx: 2 <= row_idx <= 6] = simple_new
        assert len(dao) == 0  # Insert with callable should have no effect at all
        populate(dao, 5)
        dao[lambda row_idx: 4 <= row_idx <= 6] = simple_new
        for i in range(4, 7):
            simple_at = dao[i]
            assert simple_new.id == simple_at.id
            assert simple_new.name == simple_at.name

    def test_len(self, dao):
        populate(dao)
        assert len(dao) == 1

    def test_delitem(self, dao):
        simple = populate(dao)[0]
        assert simple.row_idx in dao
        dao.delete({DAO.COL_ROW_IDX: simple.row_idx})
        assert simple.row_idx not in dao

    def test_delitem_with_callable(self, dao):
        num_row = 5
        populate(dao, num_row)
        assert len(dao) == num_row
        del dao[lambda row_idx: row_idx > 1]
        assert len(dao) == 0


def populate(dao, num=1):
    data = []
    for i in range(num):
        data.append(dao.insert(Simple(i, "simple%s" % i)))
    return data


def all_format(cell, formatter):
    return all([formatter.font.name == cell.font.name,
                formatter.fill.fgColor == cell.fill.fgColor,
                formatter.border.top.style == cell.border.top.style,
                formatter.alignment.horizontal == cell.alignment.horizontal,
                formatter.number_format == cell.number_format,
                formatter.protection.hidden == cell.protection.hidden])


class Simple(Entity):
    def __init__(self, id=0, name="simple"):
        super().__init__()
        self.id = id
        self.name = name

    def from_dict(self, values):
        super().from_dict(values)
        self.id = values[SimpleDAO.COL_ID]
        self.name = values[SimpleDAO.COL_NAME]
        return self

    def to_dict(self):
        values = super().to_dict()
        values.update({SimpleDAO.COL_ID: self.id, SimpleDAO.COL_NAME: self.name})
        return values

    def __eq__(self, other):
        if isinstance(other, Simple):
            return self.row_idx == other.row_idx and self.id == other.id and self.name == other.name
        return False


class SimpleDAO(DAO):
    TABLE_NAME = "Simple"
    COL_ID = "id"
    COL_NAME = "name"

    def worksheet_name(self):
        return SimpleDAO.TABLE_NAME

    def new_entity(self):
        return Simple()

    @staticmethod
    def on_create():
        return {SimpleDAO.TABLE_NAME: [SimpleDAO.COL_ID, SimpleDAO.COL_NAME]}
